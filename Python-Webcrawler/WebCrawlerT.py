from urllib.error import URLError
from bs4 import BeautifulSoup
import urllib.request as r
import numpy as np
import pandas as pd
import openpyxl
from urllib.parse import urljoin
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook


website = "https://www.tagesschau.de/"


url = r.urlopen("https://www.tagesschau.de/")
soup = BeautifulSoup(url, 'html.parser')
# looking for " span class="teaser__headline" " to get headlines of homepage

headlines = []
shorttexts = []
ArticleDate = []
ArticleUrls = []
ArticleTime = []
Authors = []
Topline = []
FullText = []

for t in soup.findAll("span", {"class": "teaser__topline"}):
    Topline.append(t.text.strip())


for a in soup.findAll("a",  {"class": "teaser__link"}, href=True):
    ArticleUrls.append(urljoin(website, a['href']))

appended = False
StringTmp = ""
appendedMeta = False

for link in ArticleUrls:
    try:
        urlTmp = r.urlopen(link)
    except URLError as e:
        print(e)
        ArticleDate.append("No Date")
        ArticleTime.append("No Time")
        Authors.append("No Author")
        FullText.append("Not found")
        continue
    soupTmp = BeautifulSoup(urlTmp, 'html.parser')

    for date in soupTmp.findAll("p", {"class": "metatextline"}):
        listMeta = date.text.split()
        listMeta.pop()
        ArticleTime.append(listMeta.pop())
        ArticleDate.append(listMeta.pop())
        appendedMeta = True
    if not appendedMeta:
        ArticleDate.append("No Date")
        ArticleTime.append("No Time")
    appendedMeta = False
    for author in soupTmp.findAll("div", {"class": "authorline__author"}):
        Authors.append(author.text.strip())
        appended = True
    if not appended:
        Authors.append("No Author")
    appended = False
    for p in soupTmp.findAll("p", {"class": "textabsatz m-ten m-offset-one l-eight l-offset-two columns twelve"}):
        StringTmp = StringTmp + p.text
    #test this
    if len(StringTmp) == 0:
        FullText.append("Not found")
    else:
        FullText.append(StringTmp.strip())
    StringTmp = ""


for headline in soup.findAll("span", {"class": "teaser__headline"}):
    headlines.append(headline.text.strip())


for shorttext in soup.findAll("p", {"class": "teaser__shorttext"}):
    shorttexts.append(shorttext.find_next(text=True).strip())

print(len(Authors))
print(len(ArticleDate))
print(len(headlines))
print(len(shorttexts))
print(len(FullText))
print(len(ArticleTime))

df = pd.DataFrame(
    {"Headline": headlines,
     "Text": shorttexts,
     "Date": ArticleDate,
     "Time": ArticleTime,
     "Authors": Authors,
     "URL´s": ArticleUrls,
     "Teaser": Topline,
     "Full_Article": FullText
     }
)

# if no data replace overlay with replace and only use below code with blank table
# remove this from excelwriter if no Data:  ", header=False, startrow=writer.sheets['Tabelle1'].max_row"
with pd.ExcelWriter('C:/Users/stefa/Desktop/TagesschauData.xlsx', mode="a", engine="openpyxl", if_sheet_exists='overlay') as writer:
    df.to_excel(writer, sheet_name='Sheet1', header=False, startrow=writer.sheets['Sheet1'].max_row, index=False)


dfData = pd.read_excel('C:/Users/stefa/Desktop/TagesschauData.xlsx')
dfData = dfData.drop_duplicates(subset=['Headline'])
dfData.to_excel('C:/Users/stefa/Desktop/TagesschauData.xlsx', index=False)
